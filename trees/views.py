from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Tree, TreeUpdate, TreeSpecies
from .serializers import TreeSerializer, TreeUpdateSerializer, TreeSpeciesSerializer, TreeBulkCreateSerializer
from django.db.models import Count
from django.shortcuts import get_object_or_404
import csv
import io
import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile


class IsStaffOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in ('GET', 'HEAD', 'OPTIONS'):
            return True
        return request.user and request.user.is_authenticated and request.user.is_staff

class TreeViewSet(viewsets.ModelViewSet):
    queryset = Tree.objects.all()
    serializer_class = TreeSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    @action(detail=False, methods=['post'], url_path='bulk_create')
    def bulk_create(self, request):
        # only allow staff to bulk create
        if not request.user.is_staff:
            return Response({'detail': 'forbidden'}, status=status.HTTP_403_FORBIDDEN)
        # support file upload with dry_run
        dry = request.query_params.get('dry_run') in ('1', 'true', 'True')
        file = request.FILES.get('file')
        rows = []
        errors = []
        if file:
            name = getattr(file, 'name', '')
            try:
                if name.lower().endswith('.csv'):
                    content = file.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(content))
                    for idx, r in enumerate(reader, start=1):
                        rows.append(r)
                elif name.lower().endswith(('.xls', '.xlsx')):
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    headers = [c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
                    for row in ws.iter_rows(min_row=2):
                        obj = {headers[i]: (row[i].value) for i in range(len(headers))}
                        rows.append(obj)
                else:
                    return Response({'detail':'unsupported file type'}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as exc:
                return Response({'detail': 'parse error', 'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # fallback to JSON body
            rows = request.data.get('trees', [])

        # basic validation: ensure beneficiary and planting_date
        for idx, r in enumerate(rows, start=1):
            if not r.get('beneficiary'):
                errors.append({'row': idx, 'error':'missing beneficiary'})
            if not r.get('planting_date'):
                errors.append({'row': idx, 'error':'missing planting_date'})

        if errors:
            return Response({'valid': False, 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        if dry:
            return Response({'valid': True, 'rows_parsed': len(rows), 'sample': rows[:3]})

        created = []
        for r in rows:
            try:
                t = Tree.objects.create(tree_id=r.get('tree_id') or None, species_id=r.get('species'), planting_date=r.get('planting_date'), beneficiary_id=r.get('beneficiary'))
                created.append(t)
            except Exception as exc:
                errors.append({'row': r, 'error': str(exc)})

        if errors:
            return Response({'created': len(created), 'errors': errors}, status=status.HTTP_207_MULTI_STATUS)

        return Response({'created': len(created)}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        # simple aggregations
        total = Tree.objects.count()
        by_species = Tree.objects.values('species__name').annotate(count=Count('id')).order_by('-count')[:10]
        by_status = Tree.objects.values('status').annotate(count=Count('id'))
        return Response({'total': total, 'by_species': list(by_species), 'by_status': list(by_status)})

    @action(detail=False, methods=['get'], url_path='map')
    def map(self, request):
        qs = Tree.objects.exclude(latitude__isnull=True, longitude__isnull=True)
        data = [{'id': t.pk, 'lat': t.latitude, 'lng': t.longitude, 'species': str(t.species), 'status': t.status} for t in qs]
        return Response({'points': data})

class TreeUpdateViewSet(viewsets.ModelViewSet):
    queryset = TreeUpdate.objects.all()
    serializer_class = TreeUpdateSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

    @action(detail=False, methods=['post'], url_path='add')
    def add_update(self, request):
        ser = TreeUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(TreeUpdateSerializer(obj).data, status=status.HTTP_201_CREATED)

class TreeSpeciesViewSet(viewsets.ModelViewSet):
    queryset = TreeSpecies.objects.all()
    serializer_class = TreeSpeciesSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
